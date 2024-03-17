'''
Got the code from this site: https://ryangooch.github.io/Automated-Selection-Committee/
'''

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
import csv
import datetime

import warnings
warnings.filterwarnings('ignore')

team_data_df = pd.read_csv('masseyratings.csv')

# Drop columns that are unnecessary
cols_to_drop = ["W-L","&Delta;","CMP","Sort"] # columns we won't use
# Split off human ratings and computer ratings
comp_ratings = team_data_df.drop(cols_to_drop,axis=1) # computer ratings
human_ratings = team_data_df[['Team','Conf','AP','USA']] # pulls out human ratings

summary_df = team_data_df[["Team","Conf"]] # this will hold the final results

human_ratings['AP'] = human_ratings['AP'].str.strip()
human_ratings['USA'] = human_ratings['USA'].str.strip()
human_ratings = human_ratings.apply(lambda x: pd.to_numeric(x, errors='ignore'))
comp_ratings = comp_ratings.apply(lambda x: pd.to_numeric(x, errors='ignore'))

summary_df["comp_mean"] = comp_ratings.mean(axis=1,numeric_only=True)
summary_df["human_mean"] = human_ratings.mean(axis=1,numeric_only=True,skipna=True)

def rank_calc(x, y):
    if np.isnan(y) :
        return x
    else:
        return ((2 * x + y) / 3.)

summary_df["final_rank"] = np.vectorize(rank_calc)(
        summary_df["comp_mean"], summary_df["human_mean"])

summary_df.sort_values(by=['final_rank'],inplace=True)

# Using groupby to grab the auto bids
auto_bid_teams = summary_df.groupby(['Conf']).head(1)['Team'].values

# and we can use ~isin now to get at larges
at_large_teams = summary_df[~summary_df['Team'].isin(auto_bid_teams)].head(36)['Team'].values

# all 68 teams in one array
all_68 = np.append(auto_bid_teams,at_large_teams)

final_68 = summary_df[summary_df['Team'].isin(all_68)]

#add seeds
seeds = np.array([
        1,1,1,1,
        2,2,2,2,
        3,3,3,3,
        4,4,4,4,
        5,5,5,5,
        6,6,6,6,
        7,7,7,7,
        8,8,8,8,
        9,9,9,9,
        10,10,10,10,
        11,11,11,11,11,11,
        12,12,12,12,
        13,13,13,13,
        14,14,14,14,
        15,15,15,15,
        16,16,16,16,16,16
    ])


# final_68["seed"] = seeds
final_68["seed"] = final_68.index + 1

final_68.head(16)
print(final_68)

final_68.to_csv('seeds.csv', index=False)

# # Need to tell Excel which cells get which team. Because we're
# # using 'snake' method, there are specific cells corresponding
# # to each seed and rank. Easiest way to do this is to simply
# # hard code the table in, for now
# excel_placements = [
#     'C7', 'R7', 'R39', 'C39', 'C67', 'R67', 'R35', 'C35',
#     'C27', 'R27', 'R59', 'C59', 'C51', 'R51', 'R19', 'C19',
#     'C15', 'R15', 'R47', 'C47', 'C55', 'R55', 'R23', 'C23',
#     'C31', 'R31', 'R63', 'C63', 'C43', 'R43', 'R11', 'C11',
#     'C13', 'R13', 'R45', 'C45', 'C65', 'R65', 'R33', 'C33',
#     'C25', 'R25', 'Q71', 'Q73', 'D71', 'D73', # 11 seeds
#     'C49', 'R49', 'R17', 'C17', 'C21', 'R21', 'R53', 'C53',
#     'C61', 'R61', 'R29', 'C29', 'C37', 'R37', 'R69', 'C69',
#     'C41', 'R41', 'O71', 'O73', 'F71', 'F73', 'G73'
# ]
#
# # append that to our final_68 dataframe
# final_68['excel'] = excel_placements
#
# for index, row in final_68.iterrows():
#     print(row['Team'], row['excel'])
#
# wb = Workbook()
# ws = wb.active
# for index, row in final_68.iterrows():
#     ws[row['excel']] = row['Team']
#
# # get today's date, save bracket
# today = str(datetime.date.today())
# save_file = 'bracket' + today + '.xlsx'
# wb.save(save_file)