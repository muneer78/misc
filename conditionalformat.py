import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import numpy as np

# Step 1: Import CSV as DataFrame
def load_csv(file_path):
    return pd.read_csv(file_path)

# Step 2: Apply Conditional Formatting to Excel
def apply_excel_conditional_formatting(df, filename):
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Define fill colors
    high_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    low_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
        row_has_high = any(cell.value > 50 for cell in row)
        row_has_low = any(cell.value < 20 for cell in row)

        for cell in row:
            if row_has_high:
                cell.fill = high_fill
            elif row_has_low:
                cell.fill = low_fill

    writer.save()

# Step 3: Apply Conditional Formatting to PDF
def apply_pdf_conditional_formatting(df, filename):
    fig, ax = plt.subplots(figsize=(10, len(df)*0.5))  # Adjust the figsize as needed
    ax.axis('tight')
    ax.axis('off')

    # Create a 2D list to hold colors
    cell_colors = []
    for _, row in df.iterrows():
        if any(row > 50):
            cell_colors.append(['#FF9999'] * len(row))
        elif any(row < 20):
            cell_colors.append(['#99FF99'] * len(row))
        else:
            cell_colors.append(['#FFFFFF'] * len(row))

    # Create the table
    table = ax.table(cellText=df.values,
                     colLabels=df.columns,
                     cellLoc='center',
                     loc='center',
                     cellColours=cell_colors)

    table.auto_set_font_size(False)
    table.set_fontsize(12)
    table.scale(1.2, 1.2)

    plt.savefig(filename, bbox_inches='tight')

# Main function to integrate all steps
def main(csv_file_path, excel_output_path, pdf_output_path):
    # Load CSV
    df = load_csv(csv_file_path)

    # Apply conditional formatting and save to Excel
    apply_excel_conditional_formatting(df, excel_output_path)

    # Apply conditional formatting and save to PDF
    apply_pdf_conditional_formatting(df, pdf_output_path)

# Replace with your file paths
csv_file_path = 'your_file.csv'
excel_output_path = 'output.xlsx'
pdf_output_path = 'output.pdf'

# Run the main function
main(csv_file_path, excel_output_path, pdf_output_path)