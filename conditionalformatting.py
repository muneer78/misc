import openpyxl
from openpyxl.styles import PatternFill

# Create a new Excel workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Dummy data for demonstration
data = [
    ["Name", "Score"],
    ["Alice", 85],
    ["Bob", 92],
    ["Charlie", 78],
    ["David", 95],
    ["Eva", 80],
]

# Populate the sheet with data
for row in data:
    sheet.append(row)

# Specify the range of cells to apply conditional formatting
cell_range = sheet["B2:B7"]

# Define the condition (e.g., highlight cells with values greater than 90)
condition = openpyxl.formatting.rule.CellIsRule(
    operator="greaterThan",
    formula=["90"],
    stopIfTrue=True,
    fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
)

# Apply the conditional formatting to the specified range
rule = openpyxl.formatting.rule.Rule(type="cellIs", dxf=condition)
sheet.conditional_formatting.add(cell_range, rule)

# Save the workbook
workbook.save("conditional_formatting_example.xlsx")
